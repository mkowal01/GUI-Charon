import serial
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from os import urandom
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "message_logs.xlsx"

# Funkcja tworząca plik Excel, jeśli nie istnieje
def create_excel_file():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Logs"
        ws.append(["Czas odebrania danych", "Czas wysłania ACK", "Deszyfracja danych (ms)", "Odebrana wiadomość"])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Błąd podczas tworzenia pliku Excel: {e}")

# Funkcja zapisu danych do pliku Excel
def log_to_excel(received_time, ack_time, duration_ms, message):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([received_time, ack_time, duration_ms, message])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Błąd podczas zapisywania do pliku Excel: {e}")

def load_key_from_json(index: int, json_file: str) -> bytes:
    with open(json_file, "r") as file:
        half_keys = json.load(file)
    return bytes.fromhex(half_keys[str(index)])

def xor_iv_with_index(iv: bytes, index2: bytes) -> bytes:
    index_bytes = index2[:len(iv)]
    return bytes(a ^ b for a, b in zip(iv, index_bytes))

def decrypt_data(message: bytes, json_file: str) -> str:
    index1 = message[0]  # Indeks pierwszej połowy klucza
    index2 = message[1]  # Indeks drugiej połowy klucza
    iv = message[2:14]  # IV (12 bajtów dla AES-GCM)
    tag = message[30:]  # Tag (16 bajtów dla AES-GCM)
    ciphertext = message[14:30]  # Zaszyfrowane dane

    iv = xor_iv_with_index(iv, load_key_from_json(index2, json_file))

    key1 = load_key_from_json(index1, json_file)
    key2 = load_key_from_json(index2, json_file)
    full_key = key1 + key2

    aesgcm = AESGCM(full_key)

    try:
        plaintext = aesgcm.decrypt(iv, ciphertext + tag, None)
    except Exception as e:
        raise ValueError(f"Błąd deszyfrowania: {e}")

    return plaintext.decode("utf-8")

def lora_server(port: str, baudrate: int, json_file: str):
    create_excel_file()

    try:
        with serial.Serial(port, baudrate, timeout=1) as ser:
            print(f"Serwer nasłuchuje na porcie {port} z prędkością {baudrate}.")

            while True:
                encrypted_message = ser.read(4096)
                if encrypted_message:
                    received_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
                    print(f"Otrzymano zaszyfrowaną wiadomość: {encrypted_message.hex()}")

                    try:
                        start_time = datetime.now()
                        decrypted_data = decrypt_data(encrypted_message, json_file)
                        end_time = datetime.now()

                        duration = (end_time - start_time).total_seconds() * 1000  # Czas w ms

                        print(f"Odszyfrowane dane: {decrypted_data}")

                        ack_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
                        ser.write(ack_time.encode('utf-8'))
                        print(f"Wysłano ACK z godziną: {ack_time}")

                        log_to_excel(received_time, ack_time, duration, decrypted_data)

                        if decrypted_data.strip().upper() == "TX SOCKET!!!":
                            return "SOCKET"
                    except Exception as e:
                        error_message = f"Błąd: {e}"
                        print(error_message)
                        ser.write(error_message.encode('utf-8'))
    except KeyboardInterrupt:
        print("Serwer został zatrzymany.")
    except Exception as e:
        print(f"Błąd serwera: {e}")

def main():
    json_file_path = "C:/Users/ninja/OneDrive/Pulpit/Praca/GUI-Charon/charon_library/half_keys_indexed.json"
    active_server = "LORA"  # Domyślny serwer to LoRa

    while True:
        if active_server == "LORA":
            result = lora_server("COM6", 9600, json_file_path)
            print(result)

if __name__ == "__main__":
    main()
